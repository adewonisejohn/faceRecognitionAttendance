from tkinter import messagebox

from openpyxl import load_workbook
import os
from datetime import date
from openpyxl.workbook import Workbook

class AttendanceController:
    def __init__(self):
        print("Attendance controller initialized")

    @staticmethod
    def markAttendance(attendance_data, course_code):
        # Create the directory if it doesn't exist
        folder_path = "./attendanceList"
        os.makedirs(folder_path, exist_ok=True)
        course_code = course_code.replace(" ", "-")

        todaysDate = date.today()

        # Check if the file already exists
        file_name = f"{course_code.lower()}-"+str(todaysDate)+"attendance.xlsx"
        file_path = os.path.join(folder_path, file_name)
        if os.path.exists(file_path):
            print(f"Attendance Excel file already exists at: {file_path}")
            AttendanceController.addAttendanceToExistingFile(file_path, attendance_data)
            return

        # Create a new Workbook object if the file doesn't exist
        wb = Workbook()

        # Create a new worksheet
        ws = wb.active
        ws.title = course_code.lower()

        # Add headers to the worksheet
        headers = ["First Name", "Last Name", "Matric Number", "Department", "Level"]
        ws.append(headers)

        # Add attendance data to the worksheet
        for row in attendance_data:
            ws.append(row)

        # Save the workbook to a file
        wb.save(file_path)

        print(f"Attendance Excel file created at: {file_path}")

    @staticmethod
    def addAttendanceToExistingFile(file_path, attendance_data):
        # Load the existing workbook
        wb = load_workbook(file_path)

        # Select the active worksheet
        ws = wb.active

        # Add attendance data to the worksheet
        for row in attendance_data:
            ws.append(row)

        # Save the workbook back to the file
        wb.save(file_path)

        print(f"Attendance data added to existing Excel file at: {file_path}")

# Example usage
